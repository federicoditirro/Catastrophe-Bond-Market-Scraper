# -*- coding: utf-8 -*-
"""
Created by Federico Di Tirro; 
@author: wb593691
Last Version: 06/01/2024


Instructions for use: 
This is a code to scrape Catastrophe Bond Transaction data from the Artemis Directory (https://www.artemis.bm/deal-directory/).
The code will open Chrome browser, scrape the Artemis deal directory for a specified number of transactions (see voice "links" below) and return output in an excel file called Artemis_Scraper 
The output excel file will contain data that can be extracted from the transactions on Artemis. 
If a file with the same name already exists, it will overwrite the information contained in the file, otherwise it will create a new file.
Make sure the excel file is not open while code is running, or it will return a permission error.


"""

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from dateutil.relativedelta import relativedelta
from collections import OrderedDict

import os
import re
import datetime



# Specify the path to the new working directory & Change the current working directory
directory = r"C:\Users\wb593691\Downloads" 
os.chdir(directory)
print(f"Current working directory: {os.getcwd()}")


# Create New Excel File to Store info (or overwrite existing one)
filename = "Transactions_Chart.xlsx"
sheet_name = "Transactions"


#Define the Advanced Functions Needed To Scrape Information from Description
def format_size(text):
    
    
    if "Size:" in text:
        # Extract the value after "Size:"
        size_value = text.split("Size:")[1].strip()

        # Check for currency symbol
        currency_symbol = ''
        if size_value.startswith(('$', '€', '£')):
            currency_symbol = size_value[0]
            size_value = size_value[1:]

        # Convert to numeric value
        if size_value.endswith('m'):
            numeric_value = float(size_value[:-1]) * 1e6
        elif size_value.endswith('b'):
            numeric_value = float(size_value[:-1]) * 1e9
        else:
            # Handle the case where value is just a number without 'm' or 'b'
            numeric_value = float(size_value)

        # Format the value to avoid scientific notation
        formatted_value = f"{currency_symbol}{numeric_value:,.2f}"
        return formatted_value
    else:
        return "Size not found" 

def parse_attachment_probability(description):
    # Regular expression to find "attachment probability of x%"
    probability_pattern = re.compile(r"attachment probability of (\d+(\.\d+)?)%",
                                     re.IGNORECASE)

    probability_match = probability_pattern.search(description)

    if probability_match:
        # Extract and return the probability value
        return float(probability_match.group(1))
    else:
        # Return a default or indicative value (e.g., None) if not found
        return "None"

def parse_attachment_point(description):
    # Regular expression to find "attachment point of x% of losses"
    probability_pattern = re.compile(
        r"attachment point.*?([\$€£]?)\s*(\d+(\.\d+)?)( million| billion)? of losses",
        re.IGNORECASE)

    probability_match = probability_pattern.search(description)

    if probability_match:
        # Extract the currency symbol, numeric value, and the scale (million or billion)
        currency_symbol = probability_match.group(1)
        numeric_value = probability_match.group(2)
        scale = probability_match.group(4)

        # Convert the numeric value to float
        value = float(numeric_value)

        # Adjust the value based on the scale
        if scale:
            if "million" in scale:
                value *= 1e6
            elif "billion" in scale:
                value *= 1e9

        # Format the value to avoid scientific notation and prepend the currency symbol
        formatted_value = f"{currency_symbol}{value:,.2f}"

        return formatted_value
    else:
        # Return a default or indicative value (e.g., None) if not found
        return "Unknown"

def parse_spread(description):
    # Regex patterns to capture various phrases for spread information
    spread_patterns = [
        re.compile(r"(?:spread|coupon|risk margin)(?:\s*\-?\s*equivalent)?\s*(?:to|of\s+)?(\d+(?:\.\d+)?)%", re.IGNORECASE),
        re.compile(r"(?:spread|coupon|risk margin) to be paid to investors is (\d+(?:\.\d+)?)%", re.IGNORECASE),
        re.compile(r"(?:spread|coupon|risk margin)\s+(?:fixed\s*)?(?:at|of\s+)?(\d+(?:\.\d+)?)%", re.IGNORECASE),
        re.compile(r"(\d+(?:\.\d+)?)% (?:spread|coupon|risk margin)", re.IGNORECASE),
        re.compile(r"(?:priced|settle[d]?|finali[sz]ed|fixed)\s+.*?(\d+(?:\.\d+)?)\s*%", re.IGNORECASE | re.DOTALL),
        re.compile(r"guidance(?:,)? (?:at|of) (\d+(?:,\d+)?(?:\.\d+)?)%", re.IGNORECASE),
        re.compile(r"(?:just)?\s*(above|below)\s*the\s*(?:initial|final)?\s*mid-?point\s*(?:at|of)\s*(\d+(?:\.\d+)?)%", re.IGNORECASE),
        re.compile(r"pricing (?:at|of) (\d+(?:\.\d+)?)%", re.IGNORECASE), 
        re.compile(r"(?:spread|coupon|risk margin) (:?level\s*)(?:at|of) (\d+(?:\.\d+)?)%", re.IGNORECASE),
        re.compile(r"(?:settling|pricing|spread|coupon|risk margin)\s*(?:fixed|settled|finalized|determined)?\s*(?:\sat)?(?:\sthe)?(?:\s(?:raised|lowered))?\s*(?:level)?(?:\sat|\sof)?\s*(\d+(?:\.\d+)?)%", re.IGNORECASE),

        # Other patterns specifically for basis points
        re.compile(r"(?:spread|coupon|risk margin)(?:\s*\-?\s*equivalent)?\s*(?:to|of\s+)?(\d+(?:,\d+)?(?:\.\d+)?)\s*(bps|basis points)", re.IGNORECASE),
        re.compile(r"(?:spread|coupon|risk margin) to be paid to investors is (\d+(?:,\d+)?(?:\.\d+)?)\s*(bps|basis points)", re.IGNORECASE), 
        re.compile(r"(?:priced|settle?d|finali[sz]ed|fixed)\s+.*?(\d+(?:\.\d+)?)\s*(bps|basis points)", re.IGNORECASE | re.DOTALL),
        re.compile(r"guidance(?:,)? (?:at|of) (\d+(?:,\d+)?(?:\.\d+)?)\s*(bps|basis points)", re.IGNORECASE),
        re.compile(r"(?:just)?\s*(above|below)\s*the\s*(?:initial|final)?\s*mid-?point\s*(?:at|of)\s*(\d+)\s*(bps|basis points)", re.IGNORECASE),
        re.compile(r"pricing (?:at|of) (\d+)\s*(bps|basis points)", re.IGNORECASE),
        re.compile(r"(?:spread|coupon|risk margin)\s+(?:level\s+)?(?:at|of)\s+(\d+)\s*(bps|basis points)?", re.IGNORECASE),
        re.compile(r"(?:SOFR|LIBOR)\s*(?:\+|plus)?\s*(\d+(?:\.\d+)?)\s*(bps|basis points)?", re.IGNORECASE),
        re.compile(r"(?:settling|pricing)\s+(?:remained\s+)?(?:fixed\s+)?at\s+?the\s+(?:raised|lowered)\s+(?:(bps|basis points))?", re.IGNORECASE),
        re.compile(r"(?:settling|pricing|spread|coupon|risk margin)\s*(?:remained\s+)?(?:fixed|settled|finalized|determined)?\s*(?:\sat)?(?:\sthe)?(?:\s(?:raised|lowered))?\s*(?:level)?(?:\sat|\sof)?\s*(?:(bps|basis points))?", re.IGNORECASE),

        # Other pattern
        re.compile(r"(\d+(?:\.\d+)?)% rate-on-line", re.IGNORECASE),
        re.compile(r"(\d+(?:\.\d+)?)%\s+(coupon|spread|risk margin)", re.IGNORECASE),
    ]
    
    
    # List to store rates and their positions
    matches = []

    # Check all patterns and store results
    for i, pattern in enumerate(spread_patterns):
        for match in pattern.finditer(description):
            groups = match.groups()
            rate = groups[0]
            unit = groups[1] if len(groups) > 1 and groups[1] is not None else None
            
            if rate:
                rate = float(rate.replace(',', ''))
                if unit in ("bps", "basis points"):
                    rate /= 100  # Convert basis points to percentages
                elif rate > 100:
                    rate /= 100  # Normalize rates that are presumably not in basis points but are too high
                
                # Store match with its start position and pattern index
                matches.append((rate, match.start(), i))

    # Sort matches by position and then by pattern priority
    matches.sort(key=lambda x: (x[1], -x[2]))

    # Prioritize the matches based on their pattern index (keeping specific indices as priority if needed)
    prioritized = [m for m in matches if m[2] in (0, 11, 17, 20, 21)]
    if prioritized:
        return prioritized[-1][0]  # Return the last match from the prioritized list

    # If no prioritized matches, return the last match found
    return matches[-1][0] if matches else "NA"

def check_multiple_tranche(description):
    if "tranches" in description.lower():
        return "Yes"
    else:
        return "No"

def parse_expected_loss(description):
    # Regular expression to match the required phrases and capture the expected loss value
    expected_loss_pattern = re.compile(
        r"expected loss\s*(?:\w+\s*){0,3}(?:was\s*|is\s*)?(?:set\s*at\s*|of\s*)?(?:\w+\s*){0,5}(\d+(\.\d+)?)(?:\s*%|\s*basis points|\s*bps)",
        re.IGNORECASE
    )

    expected_loss_match = expected_loss_pattern.search(description)

    if expected_loss_match:
        # Extract the expected loss value and convert it to a float
        value = float(expected_loss_match.group(1))
        unit = expected_loss_match.group(0).lower()  # Check the text within the match for unit
        if 'basis points' in unit or 'bps' in unit:
            value /= 100  # Convert basis points to percentage
        return value
    else:
        # Return a default or indicative value (e.g., None) if not found
        return "NA"

def parse_maturity(description, date_of_issue):

    def word_to_number(word):
        mapping = {
            "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
            "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
            # ... Add more mappings as needed
        }
        if word.startswith('almost '):
            number_word = word.split(' ')[1]
            return mapping.get(number_word.lower()) - 1 if mapping.get(number_word.lower()) else None
        else:
            return mapping.get(word.lower())

    def extract_date(date_string):
        try:
            return datetime.datetime.strptime(date_string, "%B %d %Y")
        except ValueError:
            # If above fails, it means the date may be in "Month Year" format
            try:
                return datetime.datetime.strptime(date_string, "%B %Y")
            except ValueError as e:
                print(f"Error parsing date '{date_string}': {e}")
                return None

    # Check for explicit start and end dates
    maturity_pattern = re.compile(
        r"maturity due in (\w+ \d{4})",
        re.IGNORECASE
    )

    start_pattern = re.compile(
        r"starting from (\w+ \d{4})",
        re.IGNORECASE
    )

    maturity_match = maturity_pattern.search(description)
    start_match = start_pattern.search(description)

    if maturity_match and start_match:
        maturity_date = extract_date(maturity_match.group(1))
        start_date = extract_date(start_match.group(1))

        if maturity_date and start_date:
            delta = relativedelta(maturity_date, start_date)
            total_years = delta.years + delta.months / 12 + delta.days / 365.25
            return round(total_years, 2)

    # Additional pattern for "over a three year term running from March 1st"
    additional_pattern = re.compile(
        r"over a (\d+|\b(?:one|two|three|four|five|six|seven|eight|nine|ten)\b) year term running from (\w+ \d{1,2})(st|nd|rd|th)?",
        re.IGNORECASE
    )

    additional_match = additional_pattern.search(description)

    if additional_match:
        period_value = additional_match.group(1)
        start_date_str = additional_match.group(2)
        current_year = datetime.datetime.now().year
        start_date = extract_date(f"{start_date_str} {current_year}")

        if start_date:
            if period_value.isalpha():
                period_value = word_to_number(period_value)
                if period_value is None:
                    return "Invalid period value"

            if isinstance(period_value, str):
                try:
                    period_value = int(period_value)
                except ValueError:
                    return "Invalid period value"

            maturity_date = start_date + relativedelta(years=int(period_value))
            delta = relativedelta(maturity_date, start_date)
            total_years = delta.years + delta.months / 12 + delta.days / 365.25
            return round(total_years, 2)

    # Original maturity logic based on issue date and period
    try:
        issue_date = datetime.datetime.strptime(date_of_issue, "%B %Y")
    except ValueError as e:
        print(f"Error parsing date '{date_of_issue}': {e}")
        return "Invalid issue date"

    period_pattern = re.compile(
        r"(?:for|of|across|to the end of|term, being on-risk until the end of)?\s*"
        r"((?:almost )?(?:\d+|\b(?:one|two|three|four|five|six|seven|eight|nine|ten)\b))\s*"
        r"(years?|months?|year|month|calendar year term)(?: term| source)?(?: of protection)?"
        r"(?: to the end of)?(?:.*?end of (\w+ \d{4}))?",
        re.IGNORECASE
    )

    period_match = period_pattern.search(description)

    if period_match:
        period_value = period_match.group(1).strip()
        period_unit = period_match.group(2).strip()
        end_date_match = period_match.group(3)

        if period_value.isalpha():
            period_value = word_to_number(period_value)
            if period_value is None:
                return "Invalid period value"

        if isinstance(period_value, str):
            try:
                period_value = int(period_value)
            except ValueError:
                return "Invalid period value"

        if end_date_match:
            try:
                end_date = datetime.datetime.strptime(end_date_match, "%B %Y")
            except ValueError:
                return "Invalid pff"
            if end_date < issue_date:
                return "Invalid maturity range"
            delta = relativedelta(end_date, issue_date)
            total_years = delta.years + delta.months / 12
            return round(total_years, 2)
        else:
            if 'month' in period_unit.lower():
                return round(period_value / 12, 2)
            else:
                return period_value
    else:
        return "Unknown"

def parse_tranche_details(description):
    # Pattern to match the tranche names and any following text until the next tranche name
    pattern = re.compile(
        r"Class\s+(?!of\b|es of\b)([A-Z0-9a-z](?:[A-Z0-9\-]*[A-Z0-9a-z])?(?![a-z]{2}))((?:(?!Class\s+of|Classes\s+of).)*?)(?=Class\s+[A-Z0-9a-z](?:[A-Z0-9\-]*[A-Z0-9a-z])?(?![a-z]{2})|$)",
        re.IGNORECASE | re.DOTALL)
    matches = pattern.findall(description)
    
    tranche_details = OrderedDict()
    
    # Iterate through all matches to gather texts for each tranche
    for tranche_name, detail_text in matches:
        if tranche_name not in tranche_details:
            tranche_details[tranche_name] = detail_text.strip()
        else:
            # Append additional text if the tranche is mentioned again
            tranche_details[tranche_name] += " " + detail_text.strip()
            

 # Post-processing to filter out overarching categories when specific subtranches are present
    final_tranche_names = list(tranche_details.keys())
    for tranche_name in list(tranche_details.keys()):
        subtranche_prefixes = [name for name in final_tranche_names if name.startswith(tranche_name + "-")]
        if subtranche_prefixes:
            final_tranche_names.discard(tranche_name)  # Remove the overarching category
    
    # Parsing details for the filtered tranches
    parsed_tranches = []
    for tranche_name in final_tranche_names:
        details = tranche_details[tranche_name]
        tranche_info = {
            'name': tranche_name,
            'attachment_probability': parse_attachment_probability(details),
            'expected_loss': parse_expected_loss(details),
            'spread': parse_spread(details),
            'attachment_point': parse_attachment_point(details),
            'tranche_description': details
        }
        parsed_tranches.append(tranche_info)
    parsed_tranches.reverse()

        
    return parsed_tranches

def find_tranche_sequence(description, total_size_million, num_tranches, tolerance=0.1):
    # Regular expression pattern to find monetary values mentioned in millions or billions
    pattern = re.compile(r'[\$€£]\s*([\d,]+(?:\.\d+)?)\s*(million|billion|m|b)', re.IGNORECASE)
    
    amounts = []
    for match in pattern.finditer(description):
        value = float(match.group(1).replace(',', ''))
        unit = match.group(2).lower()
        if unit in ['million', 'm']:
            value *= 1
        elif unit in ['billion', 'b']:
            value *= 1e3  # Convert billions to millions
        if value <= total_size_million + tolerance:
            amounts.append(value)
    
    amounts.reverse()  # Reverse the list to start matching from the bottom up
  
    def attempt_sequence(remaining_size, counts, current_sequence=[], index=0):
        # Check if a valid sequence has been found
        if counts == 0 and abs(remaining_size) <= tolerance:
            return current_sequence
        if counts <= 0 or index >= len(amounts):
            return None
        
        # Start attempting to match amounts starting from the given index
        for i in range(index, len(amounts)):
            amount = amounts[i]
            if counts == 1 and abs(remaining_size - amount) <= tolerance:
                return current_sequence + [amount]
            elif amount <= remaining_size:
                # Recursively attempt to find the rest of the sequence
                new_remaining_size = remaining_size - amount
                result = attempt_sequence(new_remaining_size, counts - 1, current_sequence + [amount], i + 1)
                if result is not None:
                    return result

        return None  # Return None if no valid sequence is found

    # Since scraping is done bottom-up, reverse the list before returning it
    sequence = attempt_sequence(total_size_million, num_tranches)

    if sequence:
        return sequence
    return "NA"


#Configure Chrome options for headless browsing and set-up Artemis URL
URL = "https://www.artemis.bm/deal-directory/"
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)
driver.get(URL)

# Initialize Last Deal
last_deal_name = None
original_last_row = None

# Headers for the sheet
headers = ["Deal", "Date of issue", "Issuer", "Sponsor", 
           "Placement / structuring agent/s", 
           "Risk modelling / calculation agents", 
           "Risks / perils covered", "Size", "Trigger type", "Ratings", 
           "Maturity", "Attachment Probability", "Attachment Point", "Multiple Tranche", 
           "Expected Loss", "Spread", "Risk Multiple", "Deal Closed", "IBRD", "Description", 
           "Link"]

# Check if the file exists and load it; otherwise, create a new workbook & add headers 
if os.path.exists(filename):
    wb = load_workbook(filename)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in reversed(range(1, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            deal_closed = ws.cell(row=row, column=18).value  # Assuming column R is the 18th column
            if cell_value and deal_closed == 1:
                last_deal_name = re.split(r'\s+Class', cell_value.strip())[0]
                original_last_row = row
                break
        # If no deals are found, set original_last_row to the row after the headers
        if original_last_row is None:
            original_last_row = 1
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        original_last_row = 1  # Set the original last row to the header row
    print("Last Closed Deal Found: ", last_deal_name, ", at row: ", original_last_row)

else:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    last_deal_name = None
    original_last_row = 1  # Header row
    print("New File Created")


# Flag to stop scraping
stop_scraping = False

    
    
#Scrape links for each deal

#Start by Retrieving Deal List and checking for new deals
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')
table = soup.find_all("table", id = "table-deal")[0]
deals = table.find_all("tr")
links = []
deal_closed_status = [] 


for deal in deals:
    tds = deal.find_all("td")
    if tds:
        Deal_name = tds[0].text.strip()
        link = tds[0].find("a").get("href")
        links.append(link)
        if 'background: #C8E6C9' in deal.get('style', ''):
            deal_closed_status.append(0)
        else:
            deal_closed_status.append(1)                                           # Deal is closed or no specific indication it's open

first_transaction = True  # Flag to indicate the first transaction
#Start scraping Links
links=links[:1000]                                                                # USE THE FOLLOWING TO CHANGE NUMBER OF TRANSACTIONS SCRAPED
for index, link in enumerate(links):
    if stop_scraping:
        break  # Break the loop if the stop flag is set    
    try:
        driver.get(link)
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
          
        try:
            # Try to extract Deal Name from the Title
            Deal_name = soup.find("div", id="info-box").find("h2").text[:-14].strip()
        except AttributeError:
            # If either the div or h2 is not found, set Deal_name to "NA"
            Deal_name = "NA"
        print(Deal_name)

        if Deal_name == last_deal_name:
            print("Matching deal found. Stopping scraping.")
            stop_scraping = True  # Set the flag to stop scraping
            break  # Break the loop if a matching deal name is found
            
        # Collect all the informaftion in the bullet points from <li> tags in a list
        data_texts = [data.text for data in soup.find_all("li")]
        deal_info = data_texts#[-100:]                                              # Use the [-100:] to limit the number of characters retrieved for ease of running code
    
        for text in deal_info:
            if "Issuer:" in text:
                Issuer = text.split("Issuer:")[1].strip()
        
            if "Cedent / sponsor: " in text:
                Sponsor= text.split("Cedent / sponsor: ")[1].strip()
                
            if "Placement / structuring agent/s:" in text:
                Placement_Structuring_agents=text.split("Placement / structuring agent/s:")[1].strip()
            
            if "Risk modelling / calculation agents etc:" in text:
                Risk_modelling_calculation_agents = text.split("Risk modelling / calculation agents etc:")[1].strip()
            
            if "Risks / perils covered:" in text:
                Risks_perils_covered = text.split("Risks / perils covered:")[1].strip()
                print("Risk Peril:",Risks_perils_covered)
            
            if "Size:" in text:
                # Extract the value after "Size:
                size_value = text.split("Size:")[1].strip()
                print("Size:",size_value)
                # Check for "Not Issued"
                if "Not" in size_value:
                    Size = "Not Issued"
                else:
                    # Initialize currency symbol
                    currency_symbol = ''
                    # Check and assign known currency symbols
                    if size_value.startswith(('$', '€', '£', 'NZ$', 'A$', 'C$')):
                        if size_value.startswith(('A$', 'C$')):
                            currency_symbol = size_value[:2]  # Capture the three-character currency symbols
                            size_value = size_value[2:]  # Remove the currency symbol from the size value
                        else:
                            currency_symbol = size_value[0]  # Capture the one-character currency symbols
                            size_value = size_value[1:]  # Remove the currency symbol from the size value
                    
                    # Use regular expressions to extract only numbers and decimal points
                    numeric_part = re.findall(r"[\d\.]+", size_value)
                    if numeric_part:
                        numeric_value_str = numeric_part[0]  # Take the first match which should be the number
                        try:
                            # Check for 'm' or 'b' multiplier and adjust accordingly
                            if 'm' in size_value.lower():
                                numeric_value = float(numeric_value_str) * 1e6
                            elif 'b' in size_value.lower():
                                numeric_value = float(numeric_value_str) * 1e9
                            else:
                                numeric_value = float(numeric_value_str)
                            Size = f"{currency_symbol}{numeric_value:,.2f}"
                        except ValueError:
                            Size = "Not determined"
                    else:
                        Size = "Not determined"
                
            
            if "Trigger type:" in text:
                Trigger_type = text.split("Trigger type:")[1].strip()
            
            if "Ratings:" in text:
                ratings = text.split("Ratings:")[1].strip()
            
            if "Date of issue:" in text:
                date_of_issue = text.split("Date of issue:")[1].strip()
                # Parse abbreviated month names and reformat to full month name
                date_object = datetime.datetime.strptime(date_of_issue, "%b %Y")
                date_of_issue = date_object.strftime("%B %Y")
                date_of_issue_formatted = datetime.datetime.strptime(date_of_issue, "%B %Y")
    
        
        description_div = soup.find("div", class_="pf-content")
        description = ' '.join(description_div.stripped_strings) if description_div else "Description not found"
        
        #Add details that must be parsed
        maturity = parse_maturity(description, date_of_issue)                       
        attachment_probability = parse_attachment_probability(description)
        expected_loss = parse_expected_loss(description)
        attachment_point = parse_attachment_point(description)
        spread = parse_spread(description)
        
        # Calculate Risk Multiple only if spread and expected_loss are known
        if spread != "NA" and expected_loss != "NA" and expected_loss > 0:
            risk_multiple = round(spread / expected_loss, 2)
        else:
            risk_multiple = "Unknown"
    
        #Check if IBRD Deal    
        ibrd_column_value = 1 if "IBRD" in Issuer or "International Bank for Reconstruction and Development" in Issuer or "World Bank" in Issuer else 0
    
    
        #Handle different tranches
        multiple_tranche = check_multiple_tranche(description)
        if multiple_tranche == "Yes":
            tranche_details = parse_tranche_details(description)
            
            # Attempt to find tranche sizes
            #If not determine overall size, not determined tranche sizes
            if Size.lower() == "not determined":
            # Assign 'Not determined' directly to all tranches if the original size is not determined
                for tranche in tranche_details:
                    tranche['size'] = "Not determined"
            
            elif Size.lower() == "not issued":
                for tranche in tranche_details:
                    tranche['size'] = "Not issued"
                
            else:
                total_size_cleaned = re.sub(r'[^\d.]', '', Size)
                total_size_numeric = float(total_size_cleaned) / 1e6  # Convert to millions
                tranche_sizes_sequence = find_tranche_sequence(description, total_size_numeric, len(tranche_details))
                
               
                # Check if the sizes sequence is found and if so assign sizes
                if isinstance(tranche_sizes_sequence, list):
                    # Assign sizes to each tranche
                    for tranche, size in zip(tranche_details, tranche_sizes_sequence):
                        full_size_value = size * 1_000_000                          # Convert millions to full value
                        tranche['currency'] = currency_symbol                       # Assuming currency_symbol is extracted from the original size parsing
                        tranche['size'] = f"{tranche['currency']}{full_size_value:,.2f}"
                else:
                    # If sizes sequence is "NA" or doesn't match the number of tranches, mark as "ERROR"
                    for tranche in tranche_details:
                        tranche['size'] = "ERROR"
                    
            for tranche in tranche_details:
                modified_deal_name = f"{Deal_name} Class {tranche['name']}"
                # Size should be taken directly from the tranche details if available
                Size = tranche.get('size', "Size not determined")
                attachment_probability = tranche['attachment_probability']
                expected_loss = tranche['expected_loss']
                if ("not issued" in tranche['tranche_description'].lower() or 
                    "not placed" in tranche['tranche_description'].lower() or 
                    "pulled from issuance" in tranche['tranche_description'].lower() or 
                    "won't be issued" in tranche['tranche_description'].lower() or 
                    "no longer be issued" in tranche['tranche_description'].lower() or 
                    "will not now be placed" in tranche['tranche_description'].lower()):
                    tranche['spread'] = "Not issued"
                    spread = tranche['spread']
                else:
                    spread = tranche['spread']
                    
                attachment_point = tranche['attachment_point']
                currency = Size[0] if Size not in ["Not determined", "ERROR", "NA"] else "NA"
                if Size not in ["Not issued", "ERROR", "NA"]:
                    try:
                        amount = float(re.sub(r'[^\d.]', '', Size[1:]))             # Exclude the currency symbol and convert to float
                    except ValueError:  
                        amount = "NA"                                               # Handle case where conversion fails


                #print statement for debugging:
                print(Deal_name)
                print(len(tranche_details))
                print("Tranche Name:", tranche['name'])
                print("Attachment Probability:", tranche['attachment_probability'])
                print("Expected Loss:", tranche['expected_loss'])
                print("Spread:", tranche['spread'])
                print("Attachment Point:", tranche['attachment_point'])
                print("Tranche Description Text:", [tranche['tranche_description']])
                print(f"Total size extracted: {total_size_numeric} million")
                print(f"Tranche details parsed: {len(tranche_details)} tranches found")
                tranche_sizes_sequence = find_tranche_sequence(description, total_size_numeric,len(tranche_details))
                print(f"Tranche sizes sequence: {tranche_sizes_sequence}")
                print("-------------------------------------------")
                
                # Risk Multiple calculation
                if spread != "NA" and expected_loss != "NA" and expected_loss > 0:
                    risk_multiple = round(spread / expected_loss, 2)
                else:
                    risk_multiple = "NA"
            
                # Append row data
    
                row_data = [modified_deal_name, date_of_issue_formatted, Issuer, Sponsor,
                            Placement_Structuring_agents, Risk_modelling_calculation_agents,
                            Risks_perils_covered, Size, Trigger_type, ratings, maturity,
                            attachment_probability, attachment_point, "Yes",
                            expected_loss, spread, risk_multiple, deal_closed_status[index],
                            1 if "IBRD" in Issuer or "International Bank for Reconstruction and Development" in Issuer or "World Bank" in Issuer else 0, 
                            description, link]

                ws.insert_rows(original_last_row + 1)  # Insert a new row below the original last row
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=original_last_row + 1, column=col, value=value)
                                
                                
        else:
            currency = Size[0] if Size not in ["Not determined", "ERROR", "NA"] else "NA"
            if Size not in ["Not issued", "ERROR", "NA"]:
                try:
                    amount = float(re.sub(r'[^\d.]', '', Size[1:]))  # Exclude the currency symbol and convert to float

                except ValueError:
                    amount = "NA"  # Handle case where conversion fails
        
            row_data = [Deal_name, date_of_issue_formatted, Issuer, Sponsor,
                        Placement_Structuring_agents, Risk_modelling_calculation_agents,
                        Risks_perils_covered, Size, Trigger_type, ratings, maturity,
                        attachment_probability, attachment_point, "No",
                        expected_loss, spread, risk_multiple, deal_closed_status[index],
                        1 if "IBRD" in Issuer or "International Bank for Reconstruction and Development" in Issuer or "World Bank" in Issuer else 0,
                        description, link]
            
            #Append Row Data            
            ws.insert_rows(original_last_row + 1)  # Insert a new row below the original last row
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=original_last_row + 1, column=col, value=value)
    except Exception as e:                                                       # Handle the error: log it, print it, or even write it to a file
        print(f"Error processing transaction {link}: {e}")
        continue  # Continue with the next transaction
wb.save(filename)

#Final Formatting

# Set the font of the first row to bold
bold_font = Font(bold=True)                                                    
for cell in ws["1:1"]:
    cell.font = bold_font


# make the Multiple tranche deals highlighted yellow

# Define a yellow fill style
yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

# Loop through the rows and apply yellow fill if "Multiple Tranche" is "Yes"
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    if row[7].value == "ERROR":
        for cell in row:
            cell.fill = yellow_fill

#Set a header row style
header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

# Define the font color and style for the header
header_font = Font(color="FFFFFF", bold=True)

# Apply styles to the header row
for cell in ws["1:1"]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Set the width of each column to a custom value and the 
column_width = 35                                                               # Example width, adjust as needed
for col in ws.columns:
    column = col[0].column_letter                                               # Get the column letter
    ws.column_dimensions[column].width = column_width

    
# Define the border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


# Set the height of the header row
header_row_height = 25# Adjust the height value as needed
ws.row_dimensions[1].height = header_row_height

# Apply thin border to all cells in  worksheet
for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border
        
#Apply thick bottom border to Header
for cell in ws[1]:
    cell.border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick', color="FFFFFF"))


wb.save(filename)
